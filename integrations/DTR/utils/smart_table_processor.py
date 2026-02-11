"""
Smart Table Processor - 通用表格智能处理器
自动检测并处理复杂Excel表格（多级表头、合并单元格、非标准格式等）
"""

import pandas as pd
import re
from typing import Dict, List, Tuple, Optional
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class SmartTableProcessor:
    """智能表格处理器 - 自动处理各种复杂表格格式"""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required. Install: pip install openpyxl")
    
    def process_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[pd.DataFrame, Dict]:
        """
        智能处理Excel文件
        
        Args:
            file_path: Excel文件路径
            sheet_name: 指定sheet名称（可选，None表示使用active sheet）
        
        Returns:
            (cleaned_df, metadata)
        """
        
        # 1. 检测表格结构
        structure = self._detect_structure(file_path, sheet_name=sheet_name)
        
        # 2. 处理合并单元格
        if structure['has_merged_cells']:
            df = self._process_merged_cells(file_path, sheet_name=sheet_name)
        else:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
        
        # 3. 检测并跳过表头行
        header_end_row = self._detect_header_end(df)
        if header_end_row > 0:
            df = df.iloc[header_end_row:].reset_index(drop=True)
        
        # 4. 智能提取列名
        df = self._extract_column_names(df, file_path, header_end_row, sheet_name=sheet_name)
        
        # 5. 清理数据
        df = self._clean_data(df)
        
        # 6. 生成metadata
        metadata = {
            "original_file": file_path,
            "header_rows_skipped": header_end_row,
            "has_merged_cells": structure['has_merged_cells'],
            "cleaned_shape": df.shape,
            "column_mapping": structure.get('column_mapping', {})
        }
        
        return df, metadata
    
    def _detect_structure(self, file_path: str, sheet_name: Optional[str] = None) -> Dict:
        """检测表格结构
        
        Args:
            file_path: Excel文件路径
            sheet_name: Sheet名称（可选）
        """
        
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # Fallback to active sheet
                ws = wb.active
        else:
            ws = wb.active
        
        # 检测合并单元格
        has_merged = len(list(ws.merged_cells.ranges)) > 0
        
        return {
            "has_merged_cells": has_merged,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column
        }
    
    def _process_merged_cells(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """处理合并单元格
        
        Args:
            file_path: Excel文件路径
            sheet_name: Sheet名称（可选）
        """
        
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
        else:
            ws = wb.active
        
        # 取消合并并填充值
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            
            # 获取左上角值
            top_left_value = ws.cell(min_row, min_col).value
            
            # 取消合并
            ws.unmerge_cells(str(merged_range))
            
            # 填充所有单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row, col).value = top_left_value
        
        # 保存到临时文件并读取
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_path = tmp.name
            wb.save(temp_path)
        
        df = pd.read_excel(temp_path, sheet_name=sheet_name if sheet_name else 0)
        os.unlink(temp_path)
        
        return df
    
    def _detect_header_end(self, df: pd.DataFrame) -> int:
        """
        检测表头结束位置（数据开始行）
        
        策略：
        1. 查找第一个包含大量数字的行
        2. 排除包含"Year"等关键词但值为NaN的行
        """
        
        for idx in range(min(15, len(df))):
            row = df.iloc[idx]
            
            # 统计数字列
            numeric_count = sum(1 for val in row if self._is_numeric(val))
            total_non_null = sum(1 for val in row if pd.notna(val))
            
            # 如果超过50%是数字，认为是数据行
            if total_non_null > 0 and numeric_count / total_non_null >= 0.5:
                return idx
        
        return 0
    
    def _is_numeric(self, val) -> bool:
        """判断是否为数字"""
        if pd.isna(val):
            return False
        
        if isinstance(val, (int, float)):
            return True
        
        # 尝试转换字符串
        if isinstance(val, str):
            try:
                float(val.replace(',', ''))
                return True
            except:
                return False
        
        return False
    
    def _extract_column_names(self, df: pd.DataFrame, file_path: str, header_end_row: int, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        智能提取列名 - 从多行header中合并信息
        
        Args:
            df: DataFrame
            file_path: 原始文件路径
            header_end_row: 表头结束行
            sheet_name: Sheet名称（可选）
        
        策略：
        1. 从header区域提取所有有意义的值
        2. ⭐ 保留父级分组信息（如"Civilian labor force" → "Total"）
        3. 合并多行信息形成完整列名
        """
        
        if header_end_row > 0:
            # 重新读取原始文件的header部分
            if sheet_name:
                df_raw = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df_raw = pd.read_excel(file_path)
            header_rows = df_raw.iloc[:header_end_row]
            
            new_columns = []
            for col_idx in range(len(df.columns)):
                # 从多行中收集该列的所有非空值
                col_values = []
                for row_idx in range(len(header_rows)):
                    val = header_rows.iloc[row_idx, col_idx]
                    if pd.notna(val):
                        val_str = str(val).strip()
                        # 过滤掉无意义的值
                        if val_str and val_str not in col_values and not val_str.startswith('['):
                            # 跳过顶层标题（如"HOUSEHOLD DATA"）
                            if 'HOUSEHOLD DATA' not in val_str and 'ANNUAL AVERAGES' not in val_str and 'Numbers in thousands' not in val_str:
                                col_values.append(val_str)
                
                # ⭐ 改进：合并父级和子级信息，形成完整描述
                if col_values:
                    if len(col_values) == 1:
                        # 只有一个值，直接使用
                        clean_name = self._clean_column_name(col_values[0])
                    else:
                        # 多个值：优先保留"父级 - 子级"结构
                        # 例如："Civilian labor force" + "Total" → "Civilian_labor_force_Total"
                        # 或 "Employed" + "Total" → "Employed_Total"
                        
                        # 过滤掉重复词（如果子级包含父级的关键词）
                        filtered_values = []
                        for v in col_values:
                            # 保留有意义的分组词
                            if any(kw in v.lower() for kw in ['civilian', 'labor', 'employed', 'unemployed', 'force']):
                                filtered_values.append(v)
                            elif v not in filtered_values:  # 其他词去重
                                filtered_values.append(v)
                        
                        if len(filtered_values) > 1:
                            # 合并为"父级_子级"
                            clean_name = "_".join([self._clean_column_name(v) for v in filtered_values])
                        else:
                            clean_name = self._clean_column_name(filtered_values[0] if filtered_values else col_values[-1])
                    
                    new_columns.append(clean_name)
                else:
                    new_columns.append(f"Column_{col_idx}")
            
            df.columns = new_columns
        
        # 确保列名唯一
        df.columns = self._make_unique_columns(list(df.columns))
        
        # ⭐ CRITICAL: 修复列名识别问题
        df = self._fix_column_names(df)
        
        return df
    
    def _fix_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """修复常见的列名识别错误"""
        
        if df.empty or len(df) < 3:
            return df
        
        # 1. 检测第一列是否为年份序列
        first_col = df.columns[0]
        first_col_data = df[first_col].dropna()
        
        if len(first_col_data) >= 3:
            try:
                # 尝试转换为数值
                numeric_data = pd.to_numeric(first_col_data, errors='coerce').dropna()
                
                if len(numeric_data) >= 3:
                    # 检查是否为年份范围 (1900-2100)
                    is_year_like = (
                        numeric_data.min() >= 1900 and
                        numeric_data.max() <= 2100 and
                        all(numeric_data == numeric_data.astype(int))  # 都是整数
                    )
                    
                    if is_year_like:
                        # 重命名为Year
                        print(f"  ✓ Detected year column: '{first_col}' → 'Year'")
                        df = df.rename(columns={first_col: 'Year'})
                        
            except:
                pass
        
        # 2. 修复无意义的列名 (Column_N, Unnamed, etc.)
        new_names = {}
        for i, col in enumerate(df.columns):
            # ⭐ PHASE 1 FIX: 确保列名是字符串
            col_str = str(col) if not isinstance(col, str) else col
            if col_str.startswith('Column_') or col_str.startswith('Unnamed'):
                # 尝试从数据推断列名
                col_data = df[col].dropna()
                if len(col_data) > 0:
                    # 检查是否全是数值
                    try:
                        pd.to_numeric(col_data, errors='raise')
                        new_names[col] = f'Value_{i}'
                    except:
                        new_names[col] = f'Text_{i}'
        
        if new_names:
            df = df.rename(columns=new_names)
        
        return df
    
    def _find_best_header_row(self, header_rows: pd.DataFrame) -> Optional[int]:
        """找到最佳的列名行"""
        
        keywords = ['year', 'total', 'agriculture', 'employed', 'population', 
                   'civilian', 'labor', 'industry', 'percent', 'rate']
        
        best_score = 0
        best_idx = None
        
        for idx in range(len(header_rows)):
            row = header_rows.iloc[idx]
            score = 0
            
            for val in row:
                if pd.notna(val):
                    val_str = str(val).lower()
                    # 计算包含多少关键词
                    for kw in keywords:
                        if kw in val_str:
                            score += 1
            
            if score > best_score:
                best_score = score
                best_idx = idx
        
        return best_idx
    
    def _clean_column_name(self, name: str) -> str:
        """清理列名"""
        
        # 移除换行符
        clean = name.replace('\n', ' ').replace('\r', ' ')
        
        # 合并多余空格
        clean = ' '.join(clean.split())
        
        # 移除特殊字符（保留字母、数字、空格、下划线）
        clean = re.sub(r'[^\w\s-]', '', clean)
        
        # 截断过长的名称
        if len(clean) > 50:
            clean = clean[:50]
        
        return clean.strip()
    
    def _make_unique_columns(self, columns: List[str]) -> List[str]:
        """确保列名唯一"""
        
        seen = {}
        unique_cols = []
        
        for col in columns:
            if col in seen:
                seen[col] += 1
                unique_cols.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                unique_cols.append(col)
        
        return unique_cols
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """清理数据"""
        
        # 1. 移除全空行
        df = df.dropna(how='all')
        
        # 2. 移除全空列
        df = df.dropna(axis=1, how='all')
        
        # 3. 尝试转换数字列
        for col in df.columns:
            df[col] = self._try_convert_numeric(df[col])
        
        # 4. 重置索引
        df = df.reset_index(drop=True)
        
        return df
    
    def _try_convert_numeric(self, series: pd.Series) -> pd.Series:
        """尝试将列转换为数字"""
        
        try:
            # 先移除逗号（如 "1,234" → "1234"）
            if series.dtype == 'object':
                series_clean = series.astype(str).str.replace(',', '')
                # 尝试转换
                converted = pd.to_numeric(series_clean, errors='coerce')
                
                # 如果转换后有效值比例>50%，使用转换结果
                valid_ratio = converted.notna().sum() / len(converted) if len(converted) > 0 else 0
                if valid_ratio > 0.5:
                    return converted
            
            # 否则尝试直接转换，如果失败则返回原series
            try:
                return pd.to_numeric(series, errors='raise')
            except (ValueError, TypeError):
                return series
        except:
            return series


# 测试函数
def main():
    """测试智能处理器"""
    
    processor = SmartTableProcessor()
    
    test_file = "/data/deeptabularresearch/benchmarks/realhitbench/tables/employment-table01.xlsx"
    
    if Path(test_file).exists():
        print("Processing:", test_file)
        print("=" * 70)
        
        df, metadata = processor.process_excel(test_file)
        
        print("\n✓ Processing complete!")
        print(f"  Shape: {df.shape}")
        print(f"  Header rows skipped: {metadata['header_rows_skipped']}")
        print(f"  Has merged cells: {metadata['has_merged_cells']}")
        print()
        
        print("Columns:")
        for idx, col in enumerate(df.columns, 1):
            print(f"  {idx}. {col}")
        print()
        
        print("First 10 rows:")
        print(df.head(10))
        print()
        
        print("Data types:")
        print(df.dtypes)
    else:
        print(f"Test file not found: {test_file}")


if __name__ == "__main__":
    main()

