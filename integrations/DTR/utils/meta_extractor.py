"""
Meta Information Extractor
从Excel文件中提取层级结构信息（合并单元格、多级表头等）
用于帮助LLM理解复杂的表格结构
"""

import pandas as pd
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
import re

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class MetaExtractor:
    """提取Excel表格的层级结构信息"""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for meta extraction. Install: pip install openpyxl")
    
    def _clean_value(self, value) -> str:
        """清洗单元格值"""
        if value is None:
            return ""
        text = str(value).strip()
        # 移除换行符，合并空格
        text = text.replace('\n', ' ').replace('\r', ' ')
        text = ' '.join(text.split())
        return text
    
    def _detect_merged_cells(self, worksheet) -> List[Dict[str, Any]]:
        """检测合并单元格"""
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            merged_cells.append({
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col,
                "value": self._clean_value(worksheet.cell(merged_range.min_row, merged_range.min_col).value)
            })
        return merged_cells
    
    def _detect_header_rows(self, worksheet, max_scan_rows: int = 10) -> Tuple[int, List[Dict[str, Any]]]:
        """检测表头行数和层级结构
        
        Returns:
            (header_end_row, col_headers_info)
        """
        col_headers = []
        header_end_row = 1  # 默认第一行是表头
        
        # 扫描前几行，找到数据开始的位置
        for row_idx in range(1, min(max_scan_rows + 1, worksheet.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(20, worksheet.max_column + 1)):  # 只检查前20列
                cell_value = worksheet.cell(row_idx, col_idx).value
                row_values.append(cell_value)
            
            # 判断是否是数据行（包含多个数字）
            numeric_count = sum(1 for v in row_values if isinstance(v, (int, float)))
            
            if numeric_count >= len(row_values) * 0.5 and numeric_count >= 3:
                # 认为这一行是数据行，表头到此结束
                header_end_row = row_idx - 1
                break
        
        if header_end_row < 1:
            header_end_row = 1
        
        # 提取列表头信息
        for level in range(1, header_end_row + 1):
            level_headers = []
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(level, col_idx)
                value = self._clean_value(cell.value)
                
                level_headers.append({
                    "level": level - 1,  # 0-based
                    "col": col_idx,
                    "value": value if value else f"[EMPTY_COL_{col_idx}]"
                })
            
            col_headers.append({
                "level": level - 1,
                "headers": level_headers
            })
        
        return header_end_row, col_headers
    
    def _build_hierarchy_triplets(self, col_headers: List[Dict[str, Any]], merged_cells: List[Dict[str, Any]]) -> List[str]:
        """构建层级关系的三元组
        
        Returns:
            三元组列表，如：
            - (table, has_column_header, "Year")
            - ("Employment Status", has_child, "Employed")
        """
        triplets = []
        
        # 1. 提取顶层列表头（level 0）
        if col_headers:
            top_level = col_headers[0]["headers"]
            seen_values = set()
            
            for header in top_level:
                value = header["value"]
                if value and not value.startswith("[EMPTY_") and value not in seen_values:
                    triplets.append(f'(table, has_column_header, "{value}")')
                    seen_values.add(value)
        
        # 2. 根据合并单元格推断层级关系
        # 如果一个合并单元格跨越多列，它的子列是它下面的表头
        for merged in merged_cells:
            parent_value = merged["value"]
            if not parent_value or parent_value.startswith("[EMPTY_"):
                continue
            
            # 找到这个合并单元格下方的列表头（作为子节点）
            parent_row = merged["min_row"]
            parent_col_start = merged["min_col"]
            parent_col_end = merged["max_col"]
            
            # 查找下一层级的表头
            if parent_row < len(col_headers):
                next_level = col_headers[parent_row]["headers"]  # parent_row作为索引（0-based后）
                
                for header in next_level:
                    if parent_col_start <= header["col"] <= parent_col_end:
                        child_value = header["value"]
                        if child_value and not child_value.startswith("[EMPTY_"):
                            triplets.append(f'("{parent_value}", has_child, "{child_value}")')
        
        return triplets
    
    def _unmerge_and_fill(self, worksheet) -> None:
        """处理合并单元格：取消合并并填充值"""
        merged_ranges = list(worksheet.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            # 获取合并区域的顶左单元格值
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            top_left_value = worksheet.cell(min_row, min_col).value
            
            # 取消合并
            worksheet.unmerge_cells(str(merged_range))
            
            # 填充所有单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    worksheet.cell(row, col).value = top_left_value
    
    def extract_meta_info(self, file_path: str, max_headers: int = 30, preprocess_merged: bool = True, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """从Excel文件中提取meta信息
        
        Args:
            file_path: Excel文件路径
            max_headers: 最多提取多少个列表头
            preprocess_merged: 是否预处理合并单元格（取消合并并填充）
            sheet_name: 指定sheet名称（可选，None表示使用active sheet）
            
        Returns:
            meta_info字典，包含：
            - file_path: 文件路径
            - sheet_name: Sheet名称
            - merged_cells: 合并单元格信息
            - col_headers: 列表头信息
            - hierarchy_triplets: 层级关系三元组
            - summary: 摘要文本
            - cleaned_df: 预处理后的DataFrame（如果preprocess_merged=True）
        """
        if not Path(file_path).exists():
            return {
                "error": f"File not found: {file_path}",
                "file_path": file_path,
                "sheet_name": sheet_name,
                "merged_cells": [],
                "col_headers": [],
                "hierarchy_triplets": [],
                "summary": "File not found"
            }
        
        try:
            # 加载Excel文件
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            # 选择worksheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return {
                        "error": f"Sheet '{sheet_name}' not found in file",
                        "file_path": file_path,
                        "sheet_name": sheet_name,
                        "merged_cells": [],
                        "col_headers": [],
                        "hierarchy_triplets": [],
                        "summary": f"Sheet not found. Available sheets: {workbook.sheetnames}"
                    }
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
                sheet_name = worksheet.title
            
            # 1. 检测合并单元格
            merged_cells = self._detect_merged_cells(worksheet)
            
            # 1.5 预处理合并单元格（如果需要）
            if preprocess_merged and merged_cells:
                self._unmerge_and_fill(worksheet)
            
            # 2. 检测表头层级
            header_end_row, col_headers = self._detect_header_rows(worksheet)
            
            # 3. 构建层级三元组
            hierarchy_triplets = self._build_hierarchy_triplets(col_headers, merged_cells)
            
            # 4. 提取关键信息用于摘要
            col_names = []
            if col_headers:
                for header_info in col_headers:
                    for h in header_info["headers"][:max_headers]:
                        value = h["value"]
                        if value and not value.startswith("[EMPTY_") and value not in col_names:
                            col_names.append(value)
            
            # 5. 生成摘要
            summary_lines = []
            summary_lines.append(f"Table has {len(col_headers)} header levels")
            summary_lines.append(f"Columns: {', '.join(col_names[:10])}")
            
            if merged_cells:
                summary_lines.append(f"Contains {len(merged_cells)} merged cells (complex structure)")
                # 列出主要的合并单元格
                major_merges = [m["value"] for m in merged_cells if m["value"] and not m["value"].startswith("[EMPTY_")][:5]
                if major_merges:
                    summary_lines.append(f"Major headers: {', '.join(major_merges)}")
            
            summary = "\n".join(summary_lines)
            
            # 6. 获取数据区域信息
            data_start_row = header_end_row + 1
            data_rows = worksheet.max_row - header_end_row
            
            # 7. 生成干净的DataFrame（如果预处理了合并单元格）
            cleaned_df = None
            if preprocess_merged:
                try:
                    import pandas as pd
                    import tempfile
                    import os
                    
                    # 保存处理后的工作簿到临时文件
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                        temp_path = tmp.name
                        workbook.save(temp_path)
                    
                    # 用pandas读取（指定sheet）
                    cleaned_df = pd.read_excel(temp_path, sheet_name=sheet_name if sheet_name else 0)
                    
                    # 删除临时文件
                    os.unlink(temp_path)
                except Exception as e:
                    # 如果失败，cleaned_df保持None
                    pass
            
            meta_info = {
                "file_path": file_path,
                "sheet_name": worksheet.title,
                "header_levels": len(col_headers),
                "header_end_row": header_end_row,
                "data_start_row": data_start_row,
                "total_rows": worksheet.max_row,
                "total_cols": worksheet.max_column,
                "data_rows": data_rows,
                "merged_cells": merged_cells,
                "col_header_names": col_names,
                "hierarchy_triplets": hierarchy_triplets[:15],  # 限制数量，保持简洁
                "summary": summary,
                "cleaned_df": cleaned_df  # 预处理后的DataFrame
            }
            
            return meta_info
            
        except Exception as e:
            return {
                "error": str(e),
                "file_path": file_path,
                "merged_cells": [],
                "col_headers": [],
                "hierarchy_triplets": [],
                "summary": f"Error extracting meta info: {str(e)}"
            }
    
    def format_for_llm_prompt(self, meta_info: Dict[str, Any], concise: bool = True) -> str:
        """将meta信息格式化为LLM prompt的一部分
        
        Args:
            meta_info: extract_meta_info返回的字典
            concise: 是否使用简洁格式（推荐）
            
        Returns:
            格式化的文本，可直接插入到prompt中
        """
        if "error" in meta_info:
            return f"⚠️  Table structure could not be analyzed: {meta_info['error']}"
        
        if concise:
            # 简洁格式：只提供关键信息
            lines = []
            lines.append("## Table Structure (Important):")
            
            # 基本信息
            lines.append(f"- Data starts at row {meta_info['data_start_row']} ({meta_info['data_rows']} data rows)")
            
            # 合并单元格警告
            if meta_info['merged_cells']:
                lines.append(f"- ⚠️  {len(meta_info['merged_cells'])} merged cells detected (already preprocessed)")
            
            # 关键列表头（只列出前10个）
            if meta_info['col_header_names']:
                key_cols = ', '.join(meta_info['col_header_names'][:10])
                lines.append(f"- Key columns: {key_cols}")
            
            # 层级结构（只列出前5个）
            if meta_info['hierarchy_triplets']:
                lines.append(f"- Hierarchical structure detected ({len(meta_info['hierarchy_triplets'])} relationships):")
                for triplet in meta_info['hierarchy_triplets'][:5]:
                    lines.append(f"  {triplet}")
            
            return "\n".join(lines)
        else:
            # 详细格式（原来的逻辑）
            lines = []
            lines.append("## Table Structure (Meta Information)")
            lines.append(f"- File: {Path(meta_info['file_path']).name}")
            lines.append(f"- Header Levels: {meta_info['header_levels']}")
            lines.append(f"- Total Columns: {meta_info['total_cols']}")
            lines.append(f"- Data Rows: {meta_info['data_rows']} (starting from row {meta_info['data_start_row']})")
            
            # 合并单元格信息
            if meta_info['merged_cells']:
                lines.append(f"\n⚠️  **Contains {len(meta_info['merged_cells'])} merged cells** - Complex structure!")
            
            # 列表头
            if meta_info['col_header_names']:
                lines.append(f"\n### Column Headers:")
                for idx, col_name in enumerate(meta_info['col_header_names'][:20], 1):
                    lines.append(f"  {idx}. {col_name}")
            
            # 层级关系
            if meta_info['hierarchy_triplets']:
                lines.append(f"\n### Hierarchical Relationships:")
                for triplet in meta_info['hierarchy_triplets'][:15]:
                    lines.append(f"  - {triplet}")
            
            lines.append(f"\n### Summary:")
            lines.append(meta_info['summary'])
            
            return "\n".join(lines)


# 测试函数
def main():
    """测试meta提取功能"""
    extractor = MetaExtractor()
    
    # 测试文件
    test_file = "/data/deeptabularresearch/benchmarks/realhitbench/tables/employment-table01.xlsx"
    
    if Path(test_file).exists():
        print(f"Extracting meta info from: {test_file}\n")
        meta_info = extractor.extract_meta_info(test_file)
        
        print("="*80)
        print("RAW META INFO:")
        print("="*80)
        import json
        print(json.dumps({k: v for k, v in meta_info.items() if k != 'merged_cells'}, indent=2))
        
        print("\n" + "="*80)
        print("FORMATTED FOR LLM:")
        print("="*80)
        print(extractor.format_for_llm_prompt(meta_info))
    else:
        print(f"Test file not found: {test_file}")


if __name__ == "__main__":
    main()

